"""
export_xlsx.py — workbook builders for the per-page Excel exports.

Read-only: builders take plain data and return openpyxl worksheets; no DB
access, no Flask. routes_export.py does the querying and calls in here.

Shared conventions (locked):
  • Header block on every sheet: report name / period / date basis /
    generated-at, then a blank row, then the bold table-header row.
  • Money format $#,##0.00; negatives in parens-free minus style.
  • Frozen panes just under the table-header row.
  • Bold subtotal/total rows.
  • Signs are applied by the CALLER (true cash direction: in +, out −).
"""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONEY_FMT = "$#,##0.00"
PCT_FMT = "0.00%"

_TITLE_FONT = Font(bold=True, size=13)
_META_FONT = Font(size=9, color="666666")
_HDR_FONT = Font(bold=True)
_TOTAL_FONT = Font(bold=True)
_HDR_BORDER = Border(bottom=Side(style="thin", color="999999"))


def month_label(key):
    """'2026-03' → 'Mar 2026'; '03' → 'Mar' (legacy single-year keys)."""
    MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    if "-" in key:
        y, m = key.split("-")
        return f"{MONTHS[int(m) - 1]} {y}"
    return MONTHS[int(key) - 1]


def month_labels(month_keys):
    """Per spec: show the year only when the range crosses years."""
    years = {k.split("-")[0] for k in month_keys if "-" in k}
    if len(years) <= 1:
        return [month_label(k).split(" ")[0] if "-" in k else month_label(k)
                for k in month_keys]
    return [month_label(k) for k in month_keys]


def _write_header_block(ws, title, period_label, basis_label=None):
    """Rows 1-3: title / period · basis / generated-at. Returns the row
    number where the table header should go (after one blank row)."""
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    meta = f"Period: {period_label}"
    if basis_label:
        meta += f"  ·  Date basis: {basis_label}"
    ws.cell(row=2, column=1, value=meta).font = _META_FONT
    ws.cell(row=3, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ).font = _META_FONT
    return 5  # row 4 stays blank


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_transactions_sheet(wb, sheet_name, *, title, period_label,
                             basis_label, columns, rows, total_key=None):
    """Flat transaction listing.

    columns: list of dicts {header, key, money?, width?}
    rows:    list of dicts keyed by column keys
    total_key: column key to sum into a bold bottom Total row (None → none)
    """
    ws = wb.create_sheet(sheet_name)
    hdr_row = _write_header_block(ws, title, period_label, basis_label)

    for ci, col in enumerate(columns, start=1):
        c = ws.cell(row=hdr_row, column=ci, value=col["header"])
        c.font = _HDR_FONT
        c.border = _HDR_BORDER
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    r = hdr_row
    for row in rows:
        r += 1
        for ci, col in enumerate(columns, start=1):
            v = row.get(col["key"])
            c = ws.cell(row=r, column=ci, value=v)
            if col.get("money"):
                c.number_format = MONEY_FMT

    if total_key is not None and rows:
        r += 1
        ws.cell(row=r, column=1, value="Total").font = _TOTAL_FONT
        for ci, col in enumerate(columns, start=1):
            if col["key"] == total_key:
                c = ws.cell(row=r, column=ci,
                            value=round(sum((row.get(total_key) or 0)
                                            for row in rows), 2))
                c.number_format = MONEY_FMT
                c.font = _TOTAL_FONT

    widths = []
    for col in columns:
        w = col.get("width")
        if not w:
            vals = [len(str(row.get(col["key"]) or "")) for row in rows[:200]]
            w = min(max([len(col["header"])] + vals) + 2, 40)
            w = max(w, 10)
        widths.append(w)
    _autosize(ws, widths)
    return ws


def write_pivot_sheet(wb, sheet_name, *, title, period_label, basis_label,
                      row_headers, month_keys, data_rows, include_total_row=True):
    """Category × month pivot.

    row_headers: list of label-column headers, e.g. ['L1'] or ['L1', 'L2']
    data_rows:   list of dicts {'labels': [...], 'values': {month_key: float},
                                'bold': bool (optional)}
    Emits: labels | <one col per month (spec labels)> | Total, plus a bold
    bottom Total row summing every column.
    """
    ws = wb.create_sheet(sheet_name)
    hdr_row = _write_header_block(ws, title, period_label, basis_label)

    m_labels = month_labels(month_keys)
    headers = list(row_headers) + m_labels + ["Total"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font = _HDR_FONT
        c.border = _HDR_BORDER
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=len(row_headers) + 1)

    n_lbl = len(row_headers)
    col_sums = {mk: 0.0 for mk in month_keys}
    grand = 0.0
    r = hdr_row
    for row in data_rows:
        r += 1
        for ci, lbl in enumerate(row["labels"], start=1):
            c = ws.cell(row=r, column=ci, value=lbl)
            if row.get("bold"):
                c.font = _TOTAL_FONT
        row_total = 0.0
        for mi, mk in enumerate(month_keys):
            v = round(row["values"].get(mk, 0.0) or 0.0, 2)
            c = ws.cell(row=r, column=n_lbl + 1 + mi, value=v)
            c.number_format = MONEY_FMT
            if row.get("bold"):
                c.font = _TOTAL_FONT
            row_total += v
            col_sums[mk] += v
        c = ws.cell(row=r, column=n_lbl + 1 + len(month_keys),
                    value=round(row_total, 2))
        c.number_format = MONEY_FMT
        if row.get("bold"):
            c.font = _TOTAL_FONT
        grand += row_total

    if include_total_row:
        r += 1
        ws.cell(row=r, column=1, value="Total").font = _TOTAL_FONT
        for mi, mk in enumerate(month_keys):
            c = ws.cell(row=r, column=n_lbl + 1 + mi,
                        value=round(col_sums[mk], 2))
            c.number_format = MONEY_FMT
            c.font = _TOTAL_FONT
        c = ws.cell(row=r, column=n_lbl + 1 + len(month_keys),
                    value=round(grand, 2))
        c.number_format = MONEY_FMT
        c.font = _TOTAL_FONT

    label_w = [max([len(h)] + [len(str(row["labels"][i] or ""))
                               for row in data_rows]) + 2
               for i, h in enumerate(row_headers)]
    _autosize(ws, [min(max(w, 12), 40) for w in label_w]
              + [12] * len(month_keys) + [13])
    return ws


def write_table_sheet(wb, sheet_name, *, title, period_label,
                      basis_label=None, columns, rows):
    """Generic small table (portfolio total, by-group, employer years,
    income statement lines). columns: {header, key, money?, pct?, width?};
    rows may set 'bold': True."""
    ws = wb.create_sheet(sheet_name)
    hdr_row = _write_header_block(ws, title, period_label, basis_label)

    for ci, col in enumerate(columns, start=1):
        c = ws.cell(row=hdr_row, column=ci, value=col["header"])
        c.font = _HDR_FONT
        c.border = _HDR_BORDER
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    r = hdr_row
    for row in rows:
        r += 1
        for ci, col in enumerate(columns, start=1):
            v = row.get(col["key"])
            c = ws.cell(row=r, column=ci, value=v)
            if col.get("money") and v is not None:
                c.number_format = MONEY_FMT
            if col.get("pct") and v is not None:
                c.number_format = PCT_FMT
            if row.get("bold"):
                c.font = _TOTAL_FONT

    widths = []
    for col in columns:
        w = col.get("width")
        if not w:
            vals = [len(str(row.get(col["key"]) or "")) for row in rows]
            w = min(max([len(col["header"])] + vals) + 2, 40)
            w = max(w, 10)
        widths.append(w)
    _autosize(ws, widths)
    return ws


def new_workbook():
    """Workbook with the default sheet removed (builders create their own)."""
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def workbook_response_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
